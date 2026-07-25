# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "pandas",
#     "matplotlib",
#     "openpyxl",
# ]
# ///

import os
import sys
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Using openpyxl's high-level generators for formatting structures
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image

DB_PATH = "transportation_accidents.db"
OUTPUT_IMAGE = "transportation_safety_dashboard.png"
OUTPUT_EXCEL = "transportation_safety_report.xlsx"

# Mock exposure matrix: Denotes Daily Traffic Volume Index per speed zone category
VOLUME_INDEX_MAP = {
    '25 mph': 12000, 
    '35 mph': 25000, 
    '45 mph': 40000,
    '55 mph': 65000, 
    '65 mph': 95000, 
    '70 mph': 110000,
}

def check_database_exists(path):
    """Gracefully handles missing database prerequisite step."""
    if not os.path.exists(path):
        print(f"[FATAL ERROR]: Missing critical data source file at '{path}'.", file=sys.stderr)
        print("Please ensure 'transportation_accidents.db' is inside this directory.", file=sys.stderr)
        sys.exit(1)

def main():
    # Strict validation of system file prerequisites
    check_database_exists(DB_PATH)
    
    try:
        conn = sqlite3.connect(DB_PATH)
    except sqlite3.Error as e:
        print(f"[DATABASE CONNECTION ERROR]: Could not connect. Reason: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        # =====================================================================
        # PHASE 1: DATA EXTRACTION & CORE ANALYTICAL COMPUTATIONS
        # =====================================================================
        print("--- 1. Pulling Raw Crash Data Metrics ---")
        query_raw = """
            SELECT weather_condition, speed_limit, COUNT(crash_id) as crash_count
            FROM crashes
            WHERE weather_condition IS NOT NULL AND speed_limit IS NOT NULL
            GROUP BY weather_condition, speed_limit;
        """
        df_raw = pd.read_sql_query(query_raw, conn)
        df_raw['speed_zone'] = df_raw['speed_limit'].astype(str) + " mph"

        # TASK A: Cross-tabulating Weather Conditions with Speed Limit Zones
        print("\n--- 2. Executing Weather vs Speed Zone Cross-Tabulation ---")
        df_crosstab = pd.crosstab(
            index=df_raw['weather_condition'],
            columns=df_raw['speed_zone'],
            values=df_raw['crash_count'],
            aggfunc='sum'
        ).fillna(0).astype(int)
        print(df_crosstab)

        # TASK B: Calculating Exposure-Adjusted Accident Rates
        print("\n--- 3. Computing Exposure-Adjusted Rates ---")
        df_rates = df_raw.groupby('speed_zone')['crash_count'].sum().reset_index()
        
        # Enforce structural visibility for all tracking zones, including zero-crash zones
        all_zones = pd.DataFrame({'speed_zone': list(VOLUME_INDEX_MAP.keys())})
        df_rates = pd.merge(all_zones, df_rates, on='speed_zone', how='left').fillna(0)
        df_rates['crash_count'] = df_rates['crash_count'].astype(int)
        
        # Map traffic volumes and compute the final safety rate scale index
        df_rates['daily_volume'] = df_rates['speed_zone'].map(VOLUME_INDEX_MAP).fillna(50000)
        df_rates['crash_rate_per_10k'] = (df_rates['crash_count'] / df_rates['daily_volume']) * 10000
        df_rates = df_rates.sort_values(by='crash_rate_per_10k', ascending=False)
        print(df_rates)

        # NEW TASK C: Pulling Demographic Records Directly From Stored View
        print("\n--- 4. Querying Stored View for Demographic Data ---")
        query_demo = """
            SELECT age_band, injury_severity, COUNT(person_id) as total_count
            FROM view_demographic_safety_summary
            WHERE age_band IS NOT NULL AND injury_severity IS NOT NULL
            GROUP BY age_band, injury_severity;
        """
        df_demo_raw = pd.read_sql_query(query_demo, conn)
        
        # Cross-tabulate demographic outputs for clean matrix rendering inside spreadsheet
        df_demo_matrix = pd.crosstab(
            index=df_demo_raw['age_band'],
            columns=df_demo_raw['injury_severity'],
            values=df_demo_raw['total_count'],
            aggfunc='sum'
        ).fillna(0).astype(int)
        
        # Reorder KABCO column priority structure explicitly if available
        kabco_order = [c for c in ['K', 'A', 'B', 'C', 'O'] if c in df_demo_matrix.columns]
        df_demo_matrix = df_demo_matrix[kabco_order]

        # Compute marginal totals directly inside pandas before Excel generation
        # 1. Sum horizontally across columns (K, A, B, C, O) to create a 'Total' column
        df_demo_matrix['Total'] = df_demo_matrix.sum(axis=1)

        # 2. Sum vertically down rows (Age Bands) to create a 'Total' summary row
        df_demo_matrix.loc['Total'] = df_demo_matrix.sum(axis=0)

        print("\n--- Demographic Matrix with Automated Marginal Totals ---")
        print(df_demo_matrix)

        # TASK D: Pulling Injury Severity by Autonomy Classification
        print("\n--- 5. Querying Injury Severities by Autonomy Tier ---")
        query_autonomy_injury = """
            SELECT vehicle_autonomy_tier, injury_severity, COUNT(person_id) as headcount
            FROM view_demographic_safety_summary
            GROUP BY vehicle_autonomy_tier, injury_severity;
        """
        df_av_injury = pd.read_sql_query(query_autonomy_injury, conn)
        df_av_matrix = pd.crosstab(
            index=df_av_injury['vehicle_autonomy_tier'],
            columns=df_av_injury['injury_severity'],
            values=df_av_injury['headcount'],
            aggfunc='sum'
        ).fillna(0).astype(int)

        # TASK E: Aggregating Autonomous Crashes vs Manual Crashes
        print("\n--- 6. Computing Autonomous vs Manual Crash Proportions ---")
        query_crash_modes = """
            SELECT 
                CASE 
                    WHEN vehicle_autonomy_tier IN ('Non-Motorist (No Vehicle)', 'No Driving Automation') THEN 'Manual / Non-Motorist'
                    ELSE 'Autonomous Driving Mode (SAE L1-L5)'
                END as driving_mode,
                COUNT(DISTINCT person_id) as passenger_volume
            FROM view_demographic_safety_summary
            GROUP BY driving_mode;
        """
        df_crash_modes = pd.read_sql_query(query_crash_modes, conn)

        # =====================================================================
        # PHASE 2: MATPLOTLIB GRAPHICS COMPILATION
        # =====================================================================
        print(f"\n--- 7. Rendering Extended 4-Panel Analytics Canvas -> '{OUTPUT_IMAGE}' ---")
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('Comprehensive Transportation Safety & Autonomy Systems Dashboard', fontsize=16, fontweight='bold', y=0.96)

        # Panel 1 [Top-Left]: Stacked Bar Chart of Environmental Cross-tabulation
        df_crosstab.plot(kind='bar', stacked=True, ax=axes[0, 0], colormap='viridis', edgecolor='black', alpha=0.85)
        axes[0, 0].set_title('Crash Volume: Weather Condition vs Speed Zone', fontsize=11, fontweight='bold', pad=10)
        axes[0, 0].set_ylabel('Total Crash Count', fontsize=10)
        axes[0, 0].set_xlabel('Weather Condition', fontsize=10)
        axes[0, 0].grid(axis='y', linestyle='--', alpha=0.5)
        axes[0, 0].tick_params(axis='x', rotation=0)
        axes[0, 0].legend(title='Speed Limit')

        # Panel 2 [Top-Right]: Exposure-Adjusted Crash Rates Bar Chart
        bars = axes[0, 1].bar(df_rates['speed_zone'], df_rates['crash_rate_per_10k'], color='#c62828', edgecolor='black', alpha=0.8)
        axes[0, 1].set_title('Exposure-Adjusted Risk Rate (Per 10,000 Vehicles)', fontsize=11, fontweight='bold', pad=10)
        axes[0, 1].set_ylabel('Adjusted Crash Rate Index', fontsize=10)
        axes[0, 1].grid(axis='y', linestyle='--', alpha=0.5)
        for bar in bars:
            height = bar.get_height()
            axes[0, 1].annotate(f'{height:.2f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                                xytext=(0, 3), textcoords="offset points", ha='center', va='bottom', fontsize=9, fontweight='bold')

        # Panel 3 [Bottom-Left]: NEW - Injury Severity by Autonomy Classification
        df_av_matrix.plot(kind='bar', stacked=False, ax=axes[1, 0], colormap='coolwarm', edgecolor='black', alpha=0.85)
        axes[1, 0].set_title('Injury Severities (KABCO) by Vehicle Autonomy Tier', fontsize=11, fontweight='bold', pad=10)
        axes[1, 0].set_ylabel('Count of Individuals', fontsize=10)
        axes[1, 0].set_xlabel('Autonomy System Tier', fontsize=10)
        axes[1, 0].grid(axis='y', linestyle='--', alpha=0.5)
        axes[1, 0].tick_params(axis='x', rotation=15)
        axes[1, 0].legend(title='KABCO Scale')

        # Panel 4 [Bottom-Right]: NEW - Autonomous vs Manual Crash Modes Distribution
        axes[1, 1].pie(
            df_crash_modes['passenger_volume'], 
            labels=df_crash_modes['driving_mode'], 
            autopct='%1.1f%%', 
            startangle=140, 
            colors=['#0288d1', '#7b1fa2'],
            wedgeprops={'edgecolor': 'black', 'linewidth': 1, 'antialiased': True}
        )
        axes[1, 1].set_title('Crash Distribution Share: Autonomous vs Manual Systems', fontsize=11, fontweight='bold', pad=10)

        plt.tight_layout(rect=[0, 0, 1, 0.93])
        plt.savefig(OUTPUT_IMAGE, dpi=300)
        print("Extended 4-panel visual dashboard exported cleanly.")


        # =====================================================================
        # PHASE 3: EXCEL WORKBOOK AUTO-GEN & DESIGN STYLING
        # =====================================================================
        print(f"\n--- 6. Exporting Spreadsheet Report Sheets -> '{OUTPUT_EXCEL}' ---")
        
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
            df_crosstab.to_excel(writer, sheet_name='Environmental Crosstab', index=True)
            df_rates.to_excel(writer, sheet_name='Exposure Analysis', index=False, startrow=4)
            df_demo_matrix.to_excel(writer, sheet_name='Demographic Analysis', index=True)
            
            workbook = writer.book
            
            # Base table typography elements
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Formatting configurations
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006', bold=True)
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(color='006100', bold=True)

            # KPI Block Styling Definition
            kpi_title_font = Font(name="Arial", size=9, bold=False, color="595959")
            kpi_val_font = Font(name="Arial", size=16, bold=True, color="1F4E78")
            kpi_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            kpi_border = Border(
                left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0')
            )
            
            # Build the KPI Blocks on the Exposure Analysis sheet
            ws_analysis = workbook['Exposure Analysis']
            max_row_data = ws_analysis.max_row
            
            if max_row_data < 6:
                max_row_data = 6
                
            kpi_definitions = [
                {"title": "Total Tracked Crashes", "formula": f"=SUM(B6:B{max_row_data})", "cols": ["A", "B"]},
                {"title": "Avg Risk Rate Index", "formula": f"=AVERAGE(D6:D{max_row_data})", "cols": ["C", "D"]},
                {"title": "Zero-Crash Safe Zones", "formula": f'=COUNTIF(B6:B{max_row_data}, 0)', "cols": ["E", "F"]}
            ]
            
            for kpi in kpi_definitions:
                c1, c2 = kpi["cols"]
                ws_analysis.merge_cells(f"{c1}1:{c2}1")
                ws_analysis.merge_cells(f"{c1}2:{c2}3")
                ws_analysis[f"{c1}1"] = kpi["title"]
                ws_analysis[f"{c1}2"] = kpi["formula"]
                
                for r in range(1, 4):
                    for col_letter in [c1, c2]:
                        cell = ws_analysis[f"{col_letter}{r}"]
                        cell.fill = kpi_fill
                        cell.border = kpi_border

                ws_analysis[f"{c1}1"].font = kpi_title_font
                ws_analysis[f"{c1}1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_analysis[f"{c1}2"].font = kpi_val_font
                ws_analysis[f"{c1}2"].alignment = Alignment(horizontal="center", vertical="center")
                if "AVERAGE" in kpi["formula"]: ws_analysis[f"{c1}2"].number_format = '0.00'

            # Apply structured visual properties across data-driven sheets
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.views.sheetView:
                    ws.views.sheetView[0].showGridLines = True

                header_row = 5 if sheet_name == 'Exposure Analysis' else 1
                start_data_row = header_row + 1

                for row in ws.iter_rows(min_row=header_row, max_row=header_row):
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in ws.iter_rows(min_row=start_data_row, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                            if sheet_name == 'Exposure Analysis' and cell.column == 4: cell.number_format = '0.00'
                        else:
                            cell.alignment = Alignment(horizontal="left")

                        # Style properties for total summary fields
                        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        total_font = Font(name="Arial", size=11, bold=True, color="000000")

                        if sheet_name == 'Demographic Analysis':
                            # Check if cell is in the bottom-most row or the right-most column
                            if cell.row == ws.max_row or cell.column == ws.max_column:
                                cell.fill = total_fill
                                cell.font = total_font

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col if cell.row >= header_row)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

            # Inject Conditional Formatting Rules to Exposure Analysis sheet
            high_risk_rule = CellIsRule(operator='greaterThan', formula=['5'], stopIfTrue=True, fill=red_fill, font=red_font)
            ws_analysis.conditional_formatting.add(f'D6:D{max_row_data}', high_risk_rule)

            safe_zone_rule = CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=green_fill, font=green_font)
            ws_analysis.conditional_formatting.add(f'B6:B{max_row_data}', safe_zone_rule)


            ws_visuals = workbook.create_sheet(title="Visual Analytics")
            if ws_visuals.views.sheetView:
                ws_visuals.views.sheetView[0].showGridLines = True
            img = Image(OUTPUT_IMAGE)
            ws_visuals.add_image(img, 'B2')

        print("Spreadsheet compiled successfully!")
    except Exception as error:
        print(f"[RUNTIME PIPELINE ERROR]: {error}", file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()

if __name__ == '__main__':
    main()
